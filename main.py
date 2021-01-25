from urllib.request import urlopen
from bs4 import BeautifulSoup
import ssl
import pyrematch as re
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from courses import COURSES

# Cambiar por nombre y sigla (en mayúsculas)
ramos = {"IA": "IIC2613", "BDD": "IIC2413", "Software": "IIC2143",
         "Discretas": "IIC1253", "IPre": "IIC2987"}

# Ignore SSL certificate errors
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

# Style
thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     bottom=Side(style='thin'))

# Excel
wb = Workbook()
ws = wb.active
for index, data in enumerate(["Ramos", "Cupos", "Disp", "Razón", "% Tomados"]):
    cell = ws.cell(row=2, column=index + 2, value=data)
    ws.column_dimensions[get_column_letter(index + 2)].auto_size = True

    cell.font = Font(bold=True)
    cell.border = thick_border


for row_num, (ramo, sigla) in enumerate(COURSES.items()):

    # Webscraping
    url = "http://buscacursos.uc.cl/?cxml_semestre=2021-1&cxml_sigla=" + sigla.upper() + "#resultados"

    html = urlopen(url, context=ctx).read()
    soup = BeautifulSoup(html, "html.parser")
    texto = str(soup.find_all(href="javascript:;"))

    # Regex
    patron = "cantidad_dis=!disponibles{[0-9]+}&amp;cantidad_min=!total{[0-9]+}&"
    regex = re.compile(patron)
    total_disp = 0
    total_cupos = 0
    for match in regex.finditer(texto):
        disponibles = match.group("disponibles")
        cupos = match.group("total")
        total_disp += int(disponibles)
        total_cupos += int(cupos)

    # Escribe data
    for index, data in enumerate([ramo, total_cupos, total_disp,
                                  total_disp / total_cupos,
                                  (1 - total_disp / total_cupos) * 100]):
        cell = ws.cell(row=row_num + 3, column=index + 2, value=data)
        cell.border = thin_border
        cell.column_dimensions[get_column_letter(index + 2)].auto_size = True

wb.save("courses.xlsx")
